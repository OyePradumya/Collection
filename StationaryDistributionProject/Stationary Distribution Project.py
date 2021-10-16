from tkinter import *
from tkinter import messagebox

from openpyxl import *
from tkinter import filedialog
import matplotlib
matplotlib.use("TkAgg")
from matplotlib import style
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg,NavigationToolbar2Tk
from matplotlib.figure import Figure
import matplotlib.animation as animation
import matplotlib.pyplot as plt
from openpyxl import Workbook


root = Tk()
root.title("Stationary Distribution Portal")
root.resizable(width=False, height=False)
root.geometry("1550x800")
root.configure(bg="#004680")


# this is for graph
def graph(event):
    ag1 = entry_021F.get()
    ag2 = entry_05F.get()
    ag3 = entry_031F.get()
    ag4 = entry_06F.get()
    ag5 = entry_04F.get()
    ag6 = entry_07F.get()


    LARGE_FONT = ("Verdana", 12)
    style.use("ggplot")

    f = Figure(figsize=(5, 5), dpi=100)
    a = f.add_subplot(111)
    labels = ["files", "files Left", "sheet1", "sheet1 left", "sheet1", "sheet2 Left"]
    sizes = [ag1, ag2, ag3, ag4, ag5, ag6]
    colors = ["yellow", "gold", "lightskyblue", "coral", "orange", "red"]

    a.pie(sizes, labels=labels, colors=colors, shadow=True, startangle=90)
    # plt.legend(patches,labels,loc="best")
    a.axis("equal")






    class graph1(Tk):
        def __init__(self, *args, **kwargs):
            Tk.__init__(self, *args, **kwargs)
            container = Frame(self)
            container.pack(side="top", fill="both",expand="True")

            container.grid_rowconfigure(0, weight=1)
            container.grid_columnconfigure(0, weight=1)

            self.frames = {}
            frame = Startpage(container, self)
            self.frames[Startpage] = frame
            frame.grid(row=0, column=0, sticky="nsew")
            self.show_frame(Startpage)

        def show_frame(self, cont):
            frame = self.frames[cont]
            frame.tkraise()

    class Startpage(Frame):
        def __init__(self, parent, controller):
            Frame.__init__(self, parent)

            canvas = FigureCanvasTkAgg(f, self)

            pie2 = FigureCanvasTkAgg(f, self)
            canvas.draw()
            canvas.get_tk_widget().pack(side=TOP, fill=BOTH, expand=True)

            toolbar = NavigationToolbar2Tk(canvas, self)
            toolbar.update()
            canvas._tkcanvas.pack(side=BOTTOM, fill=BOTH, expand=True)

    app = graph1()

    app.mainloop()





# this is the ending of graph
def add_to_database(event):
    listBOx_01D.insert(END, ".................")
    listBOx_01D.insert(END, "PERSONAL DETAILS")
    listBOx_01D.insert(END, ".................")
    listBOx_01D.insert(END, "Fisrt Name:{}".format(entry_01B.get()))
    listBOx_01D.insert(END, "id:{}".format(entry_03B.get()))
    listBOx_01D.insert(END, "Divison:{}".format(entry_02B.get()))
    listBOx_01D.insert(END, "Course:{}".format(entry_04B.get()))
    listBOx_01D.insert(END, "Collected:{}".format(entry_09B.get()))
    listBOx_01D.insert(END, "Files:{}".format(equipment.get()))
    listBOx_01D.insert(END, "Sheet 1:{}".format(equipment1.get()))
    listBOx_01D.insert(END, "Sheet 2:{}".format(equipment2.get()))
    dest_file = "new_work_book.xlsx"
    students_workbook = load_workbook(filename=dest_file)

    students_worksheet = students_workbook.active

    row = 2
    column = 1
    while TRUE:
        if students_worksheet.cell(row=row, column=column).value:
            row += 1
            continue
        else:
            students_worksheet.cell(row=row, column=column).value = entry_01B.get()
            column += 1
            students_worksheet.cell(row=row, column=column).value = entry_03B.get()
            column += 1
            students_worksheet.cell(row=row, column=column).value = entry_02B.get()
            column += 1
            students_worksheet.cell(row=row, column=column).value = entry_04B.get()
            column += 1
            students_worksheet.cell(row=row, column=column).value = entry_09B.get()
            column += 1
            students_worksheet.cell(row=row, column=column).value = equipment.get()
            column += 1
            students_worksheet.cell(row=row, column=column).value = equipment1.get()
            column += 1
            students_worksheet.cell(row=row, column=column).value = equipment2.get()
            break

    students_workbook.save(filename=dest_file)

    # simutaneouly update done


def clear_system(event):
    entry_01B.delete(0, 100)
    entry_03B.delete(0, 100)
    entry_02B.delete(0, 100)
    entry_04B.delete(0, 100)
    entry_05B.delete(0, 100)
    entry_09B.delete(0, 100)
    equipment.set("0")
    equipment1.set("0")
    equipment2.set("0")
    # lListbox
    listBOx_01D.delete(0, 100)


def exit_the_program(event):
    root.destroy()


def update(event):
    global key_word
    row = 2
    column = 2

    dest_file = "new_work_book.xlsx"
    students_workbook = load_workbook(filename=dest_file)
    students_worksheet = students_workbook.active
    while True:

        if students_worksheet.cell(row=row, column=column).value == key_word:
            column -= 1
            students_worksheet.cell(row=row, column=column).value = None
            column += 1
            students_worksheet.cell(row=row, column=column).value = None
            column += 1
            students_worksheet.cell(row=row, column=column).value = None
            column += 1
            students_worksheet.cell(row=row, column=column).value = None
            column += 1
            students_worksheet.cell(row=row, column=column).value = None
            column += 1
            students_worksheet.cell(row=row, column=column).value = None
            column += 1
            students_worksheet.cell(row=row, column=column).value = None
            column += 1
            students_worksheet.cell(row=row, column=column).value = None
            students_workbook.save(dest_file)
            messagebox.showinfo(title="Search RESULT", message="{} Updated".format(key_word))
            break
        else:
            row += 1
            continue

    listBOx_01D.insert(END, ".................")
    listBOx_01D.insert(END, "PERSONAL DETAILS")
    listBOx_01D.insert(END, ".................")
    listBOx_01D.insert(END, "Fisrt Name:{}".format(entry_01B.get()))
    listBOx_01D.insert(END, "id:{}".format(entry_03B.get()))
    listBOx_01D.insert(END, "Divison:{}".format(entry_02B.get()))
    listBOx_01D.insert(END, "Course:{}".format(entry_04B.get()))
    listBOx_01D.insert(END, "Collected:{}".format(entry_09B.get()))
    listBOx_01D.insert(END, "Files:{}".format(equipment.get()))
    listBOx_01D.insert(END, "Sheet 1:{}".format(equipment1.get()))
    listBOx_01D.insert(END, "Sheet 2:{}".format(equipment2.get()))

    dest_file = "new_work_book.xlsx"
    students_workbook = load_workbook(filename=dest_file)

    students_worksheet = students_workbook.active

    row = 2
    column = 1
    while TRUE:
        if students_worksheet.cell(row=row, column=column).value:
            row += 1
            continue
        else:
            students_worksheet.cell(row=row, column=column).value = entry_01B.get()
            column += 1
            students_worksheet.cell(row=row, column=column).value = entry_03B.get()
            column += 1
            students_worksheet.cell(row=row, column=column).value = entry_02B.get()
            column += 1
            students_worksheet.cell(row=row, column=column).value = entry_04B.get()
            column += 1
            students_worksheet.cell(row=row, column=column).value = entry_09B.get()
            column += 1
            students_worksheet.cell(row=row, column=column).value = equipment.get()
            column += 1
            students_worksheet.cell(row=row, column=column).value = equipment1.get()
            column += 1
            students_worksheet.cell(row=row, column=column).value = equipment2.get()
            break

    students_workbook.save(filename=dest_file)
    # simulatenouly updating admin section

    global key_word3, a, b, f, x1, x2, x3
    a = int(entry_021F.get()) + int(equipment.get())
    b = int(entry_031F.get()) + int(equipment1.get())
    f = int(entry_04F.get()) + int(equipment2.get())
    x1 = int(entry_05F.get()) - int(equipment.get())
    x2 = int(entry_06F.get()) - int(equipment1.get())
    x3 = int(entry_07F.get()) - int(equipment2.get())

    entry_021F.delete(0, END)
    entry_031F.delete(0, END)
    entry_04F.delete(0, END)
    entry_05F.delete(0, END)
    entry_06F.delete(0, END)
    entry_07F.delete(0, END)
    #messagebox.showinfo(title="Search RESULT", message="{}update success thank u".format(entry_01F))
    entry_021F.insert(0, a)
    entry_031F.insert(0, b)
    entry_04F.insert(0, f)
    entry_05F.insert(0, x1)
    entry_06F.insert(0, x2)
    entry_07F.insert(0, x3)


def search_engine(event):
    global key_word
    key_word = entry_05B.get()
    row = 2
    column = 2

    dest_file = "new_work_book.xlsx"
    students_workbook = load_workbook(filename=dest_file)
    students_worksheet = students_workbook.active
    while True:
        if students_worksheet.cell(row=row, column=column).value == key_word:
            column -= 1
            entry_01B.insert(END, students_worksheet.cell(row=row, column=column).value)
            listBOx_01D.insert(END, ".................")
            listBOx_01D.insert(END, "PERSONAL DETAILS")
            listBOx_01D.insert(END, ".................")
            listBOx_01D.insert(END, "Fisrt Name:{}".format(entry_01B.get()))
            column += 1
            entry_03B.insert(END, students_worksheet.cell(row=row, column=column).value)
            listBOx_01D.insert(END, "id:{}".format(entry_03B.get()))
            column += 1
            entry_02B.insert(END, students_worksheet.cell(row=row, column=column).value)
            listBOx_01D.insert(END, "Division:{}".format(entry_02B.get()))
            column += 1
            entry_04B.insert(END, students_worksheet.cell(row=row, column=column).value)
            listBOx_01D.insert(END, "Course:{}".format(entry_04B.get()))
            column += 1
            entry_09B.insert(END, students_worksheet.cell(row=row, column=column).value)
            listBOx_01D.insert(END, "Collected:{}".format(entry_09B.get()))
            column += 1
            equipment.set(students_worksheet.cell(row=row, column=column).value)
            listBOx_01D.insert(END, "Files:{}".format(equipment.get()))
            column += 1
            equipment1.set(students_worksheet.cell(row=row, column=column).value)
            listBOx_01D.insert(END, "Sheet 1:{}".format(equipment1.get()))
            column += 1
            equipment2.set(students_worksheet.cell(row=row, column=column).value)
            listBOx_01D.insert(END, "Sheet 2:{}".format(equipment2.get()))
            break
        else:
            row += 1
            if students_worksheet.cell(row=row, column=column).value:
                continue
            else:
                higher_row = students_worksheet.max_row
                while higher_row != row:
                    row += 1
                    break
                else:
                    messagebox.showinfo("Search Results", message="{}  CURRENTLY RECORD IS NOT FOUND IN DATABASE!! TRY TO CHECK STUDENT'S DETAILS AND ADD IN THE DATABASE".format(key_word))
                    break


def delete_from_database(event):
    row = 2
    column = 2

    dest_file = "new_work_book.xlsx"
    students_workbook = load_workbook(filename=dest_file)
    students_worksheet = students_workbook.active
    while True:

        if students_worksheet.cell(row=row, column=column).value == key_word:
            column -= 1
            students_worksheet.cell(row=row, column=column).value = None
            column += 1
            students_worksheet.cell(row=row, column=column).value = None
            column += 1
            students_worksheet.cell(row=row, column=column).value = None
            column += 1
            students_worksheet.cell(row=row, column=column).value = None
            column += 1
            students_worksheet.cell(row=row, column=column).value = None
            column += 1
            students_worksheet.cell(row=row, column=column).value = None
            column += 1
            students_worksheet.cell(row=row, column=column).value = None
            column += 1
            students_worksheet.cell(row=row, column=column).value = None
            students_workbook.save(dest_file)

            entry_01B.delete(0, 100)
            entry_03B.delete(0, 100)
            entry_02B.delete(0, 100)
            entry_04B.delete(0, 100)
            entry_05B.delete(0, 100)
            entry_09B.delete(0, 100)
            equipment.set(0)
            equipment1.set(0)
            equipment2.set(0)

            listBOx_01D.delete(0, 100)
            messagebox.showinfo(title="Search RESULT", message="{} Deleted from the database".format(key_word))
            break
        else:
            row += 1
            continue


# admin part
def logout(event):
    entry_01F.delete(0, 100)
    entry_0AF.delete(0, 100)
    entry_021F.delete(0, 100)
    entry_031F.delete(0, 100)
    entry_04F.delete(0, 100)
    entry_06F.delete(0, 100)
    entry_07F.delete(0, 100)
    entry_05F.delete(0, 100)
    messagebox.showinfo(title="logout info", message=" Logout successfull")
    return


def updateadmin(event):
    global keyword_1
    keyword_1 = entry_0AF.get()
    row = 2
    column = 2
    dest_file = "new_adminsetup_book.xlsx"
    students_workbook = load_workbook(filename=dest_file)
    students_worksheet = students_workbook.active
    while True:

        if students_worksheet.cell(row=row, column=column).value == keyword_1:
            column -= 1
            students_worksheet.cell(row=row, column=column).value = None
            column += 1
            students_worksheet.cell(row=row, column=column).value = None
            column += 1
            students_worksheet.cell(row=row, column=column).value = None
            column += 1
            students_worksheet.cell(row=row, column=column).value = None
            column += 1
            students_worksheet.cell(row=row, column=column).value = None
            column += 1
            students_worksheet.cell(row=row, column=column).value = None
            column += 1
            students_worksheet.cell(row=row, column=column).value = None
            column += 1
            students_worksheet.cell(row=row, column=column).value = None
            students_workbook.save(dest_file)
            messagebox.showinfo(title="Search RESULT", message="Updated Succesfully".format(keyword_1))
            break
        else:
            row += 1
            continue

    dest_file = "new_adminsetup_book.xlsx"
    students_workbook = load_workbook(filename=dest_file)

    students_worksheet = students_workbook.active
    global x
    row = 2
    column = 1
    while TRUE:
        if students_worksheet.cell(row=row, column=column).value:
            row += 1
            continue
        else:
            students_worksheet.cell(row=row, column=column).value = entry_01F.get()
            column += 1
            students_worksheet.cell(row=row, column=column).value = entry_0AF.get()
            column += 1

            students_worksheet.cell(row=row, column=column).value = entry_021F.get()
            column += 1
            students_worksheet.cell(row=row, column=column).value = (entry_031F.get())

            column += 1

            students_worksheet.cell(row=row, column=column).value = (entry_04F.get())

            column += 1
            students_worksheet.cell(row=row, column=column).value = entry_05F.get()
            column += 1
            students_worksheet.cell(row=row, column=column).value = entry_06F.get()
            column += 1
            students_worksheet.cell(row=row, column=column).value = entry_07F.get()

            break

    students_workbook.save(filename=dest_file)


# update is done
# it is running properly
def setas(event):
    dest_file = "new_adminsetup_book.xlsx"
    students_workbook = load_workbook(filename=dest_file)

    students_worksheet = students_workbook.active

    row = 2
    column = 1
    while TRUE:
        if students_worksheet.cell(row=row, column=column).value:
            row += 1
            continue
        else:
            students_worksheet.cell(row=row, column=column).value = entry_01F.get()
            column += 1
            students_worksheet.cell(row=row, column=column).value = entry_0AF.get()
            column += 1
            students_worksheet.cell(row=row, column=column).value = entry_021F.get()
            column += 1
            students_worksheet.cell(row=row, column=column).value = entry_031F.get()
            column += 1
            students_worksheet.cell(row=row, column=column).value = entry_04F.get()
            column += 1
            students_worksheet.cell(row=row, column=column).value = entry_05F.get()
            column += 1
            students_worksheet.cell(row=row, column=column).value = entry_06F.get()
            column += 1
            students_worksheet.cell(row=row, column=column).value = entry_07F.get()
            break

    students_workbook.save(filename=dest_file)

    return


def search_left(event):
    global key_word
    key_word = entry_0AF.get()
    row = 2
    column = 2

    dest_file = "new_adminsetup_book.xlsx"
    students_workbook = load_workbook(filename=dest_file)
    students_worksheet = students_workbook.active
    while True:
        if students_worksheet.cell(row=row, column=column).value == key_word:
            column += 1
            entry_021F.insert(END, students_worksheet.cell(row=row, column=column).value)

            column += 1
            entry_031F.insert(END, students_worksheet.cell(row=row, column=column).value)

            column += 1
            entry_04F.insert(END, students_worksheet.cell(row=row, column=column).value)

            column += 1
            entry_05F.insert(END, students_worksheet.cell(row=row, column=column).value)

            column += 1
            entry_06F.insert(END, students_worksheet.cell(row=row, column=column).value)

            column += 1
            entry_07F.insert(END, students_worksheet.cell(row=row, column=column).value)

            break
        else:
            row += 1
            if students_worksheet.cell(row=row, column=column).value:
                continue
            else:
                higher_row = students_worksheet.max_row
                while higher_row != row:
                    row += 1
                    break
                else:
                    messagebox.showinfo("Search Results", message="{} is not admin IILEGAL ACCESS!".format(entry_01F))
                    break


frame_00A = Frame(root,bg="#004680")
frame_00A.pack()

label_01A = Label(frame_00A, text="STATIONARY DISTRIBUTION PORTAL",bg="#004680",fg="WHITE",font=("Copperplate", 14))
label_01A.pack()

frame_00B = Frame(root, bg="#004680")
frame_00B.pack()

label_01B = Label(frame_00B, text="Name:", bg="#004680",fg="WHITE",font=("Copperplate", 14))
label_01B.grid(row=0, column=0, padx=(10, 10), pady=(10, 10))

entry_01B = Entry(frame_00B)
entry_01B.grid(row=0, column=1, pady=(10, 10), padx=(10, 10))

label_02B = Label(frame_00B, text="Division:", bg="#004680",fg="WHITE",font=("Copperplate", 14))
label_02B.grid(row=0, column=2, sticky=E, padx=(10, 10), pady=(10, 10))

entry_02B = Entry(frame_00B)
entry_02B.grid(row=0, column=3, pady=(10, 10), padx=(10, 10))

label_03B = Label(frame_00B, text="Id:", bg="#004680",fg="WHITE",font=("Copperplate", 14))
label_03B.grid(row=1, column=0, padx=(10, 10), pady=(10, 10))

entry_03B = Entry(frame_00B)
entry_03B.grid(row=1, column=1, pady=(10, 10), padx=(10, 10))

label_04B = Label(frame_00B, text="Course:", bg="#004680",fg="WHITE",font=("Copperplate", 14))
label_04B.grid(row=0, column=4, padx=(10, 10), pady=(10, 10))

entry_04B = Entry(frame_00B)
entry_04B.grid(row=0, column=5, pady=(10, 10), padx=(10, 10))

label_05B = Label(frame_00B, text="Keyword:", bg="#004680",fg="WHITE",font=("Copperplate", 14))
label_05B.grid(row=1, column=4, padx=(10, 10), pady=(10, 10))

entry_05B = Entry(frame_00B)
entry_05B.grid(row=1, column=5, padx=(10, 10), pady=(10, 10))

label_07B = Label(frame_00B, text="FILES:", bg="#004680",fg="WHITE",font=("Copperplate", 14))
label_07B.grid(row=2, column=0, padx=(10, 10), pady=(10, 10))
# dropdown
equipment = IntVar(frame_00B)
equipment.set("0")
e18 = OptionMenu(frame_00B, equipment, "5", "4", "3", "2", '1', "0")

e18.grid(row=2, column=1)
entry_dropdown1 = equipment.get()
# dropdown1
equipment1 = IntVar(frame_00B)
equipment1.set("0")
e18 = OptionMenu(frame_00B, equipment1, "0", "1", "2")

e18.grid(row=2, column=3)
entry_dropdown2 = equipment1.get()
# dropdown2
equipment2 = IntVar(frame_00B)
equipment2.set("0")
e18 = OptionMenu(frame_00B, equipment2, "0", "2", '1')

e18.grid(row=2, column=5)
entry_dropdown3 = equipment2.get()

label_10B = Label(frame_00B, text="SHEET 1:", bg="#004680",fg="WHITE",font=("Copperplate", 14))
label_10B.grid(row=2, column=2, padx=(10, 10), pady=(10, 10))

label_011B = Label(frame_00B, text="SHEET 2:", bg="#004680",fg="WHITE",font=("Copperplate", 14))
label_011B.grid(row=2, column=4, padx=(10, 10), pady=(10, 10))

label_08B = Label(frame_00B, text="Collected", bg="#004680",fg="WHITE",font=("Copperplate", 14))
label_08B.grid(row=1, column=2, padx=(10, 10), pady=(10, 10))

entry_09B = Entry(frame_00B)
entry_09B.grid(row=1, column=3, pady=(10, 10), padx=(10, 10))
button_06B = Button(frame_00B, text="Check",bg="#789FCF",fg="#000000",font=("Copperplate", 14))
button_06B.bind("<Button-1>", search_engine)
button_06B.grid(row=2, column=6, padx=(10, 10), pady=(10, 10))

frame_00C = Frame(root, bg="#004680")
frame_00C.pack()

button_01C = Button(frame_00C, text="Add",bg="#789FCF",fg="#000000",font=("Copperplate", 14))
button_01C.bind("<Button-1>", add_to_database)
button_01C.grid(row=0, column=0, rowspan=2, columnspan=2, sticky=W, padx=(20, 20), pady=(10, 10))

button_02C = Button(frame_00C, text="Delete",bg="#789FCF",fg="#000000",font=("Copperplate", 14))
button_02C.bind("<Button-1>", delete_from_database)
button_02C.grid(row=0, column=3, rowspan=2, columnspan=2, sticky=W, padx=(20, 20), pady=(10, 10))

button_04C = Button(frame_00C, text="Clear",bg="#789FCF",fg="#000000",font=("Copperplate", 14))
button_04C.bind("<Button-1>", clear_system)
button_04C.grid(row=0, column=9, rowspan=2, columnspan=2, sticky=W, padx=(20, 20), pady=(10, 10))

button_05C = Button(frame_00C, text="Exit",bg="#789FCF",fg="#000000",font=("Copperplate", 14))
button_05C.bind("<Button-1>", exit_the_program)
button_05C.grid(row=0, column=14, rowspan=2, columnspan=2, sticky=W, padx=(20, 20), pady=(10, 10))

button_06C = Button(frame_00C, text="Update",bg="#789FCF",fg="#000000",font=("Copperplate", 14))
button_06C.bind("<Button-1>", update)
button_06C.grid(row=0, column=12, rowspan=2, columnspan=2, sticky=W, padx=(20, 20), pady=(10, 10))

frame_00D = Frame(root, bg="WHITE")
frame_00D.pack()

listBOx_01D = Listbox(frame_00D, bg="#004680",fg="white",font=("Copperplate", 14), height=10, width=60)
listBOx_01D.pack(padx=(10, 10), pady=(10, 10))
# admin
frame_00E = Frame(root , bg="#004680")
frame_00E.pack()

label_An = Label(frame_00E, text="ADMIN DETAILS", bg="#004680",fg="WHITE",font=("Copperplate", 14))
label_An.pack()

frame_00F = Frame(root, bg="#004680")
frame_00F.pack()

label_01F = Label(frame_00F, text="Admin Name:", bg="#004680",fg="WHITE",font=("Copperplate", 14))
label_01F.grid(row=0, column=0, padx=(10, 10), pady=(10, 10))

entry_01F = Entry(frame_00F)
entry_01F.grid(row=0, column=1, pady=(10, 10), padx=(10, 10))
label_0AF = Label(frame_00F, text="Password", bg="#004680",fg="WHITE",font=("Copperplate", 14))
label_0AF.grid(row=0, column=2, padx=(10, 10), pady=(10, 10))

entry_0AF = Entry(frame_00F)

entry_0AF.grid(row=0, column=3, pady=(10, 10), padx=(10, 10))

button_04F = Button(frame_00F, text="Login",bg="#789FCF",fg="#000000",font=("Copperplate", 14))
button_04F.bind("<Button-1>", search_left)
button_04F.grid(row=0, column=4, padx=(10, 10), pady=(10, 10))

button_01F = Button(frame_00F, text="Logout",bg="#789FCF",fg="#000000",font=("Copperplate", 14))
button_01F.bind("<Button-1>", logout)
button_01F.grid(row=0, column=5, rowspan=1, columnspan=1, sticky=W, padx=(20, 20), pady=(10, 10))

button_02F = Button(frame_00F, text="Register",bg="#789FCF",fg="#000000",font=("Copperplate", 14))
button_02F.bind("<Button-1>", setas)
button_02F.grid(row=0, column=6, rowspan=1, columnspan=1, sticky=W, padx=(20, 20), pady=(10, 10))

button_03F = Button(frame_00F, text="Update Info",bg="#789FCF",fg="#000000",font=("Copperplate", 14))
button_03F.bind("<Button-1>", updateadmin)
button_03F.grid(row=0, column=7, rowspan=1, columnspan=1, sticky=W, padx=(20, 20), pady=(10, 10))

label_02F = Label(frame_00F, text="Files:", bg="#004680",fg="WHITE",font=("Copperplate", 14))
label_02F.grid(row=1, column=0, padx=(10, 10), pady=(10, 10))

entry_021F = Entry(frame_00F)
entry_021F.grid(row=1, column=1, pady=(10, 10), padx=(10, 10))
entry_021F.insert(0, 0)
label_03F = Label(frame_00F, text="Sheet 1:", bg="#004680",fg="WHITE",font=("Copperplate", 14))
label_03F.grid(row=1, column=2, padx=(10, 10), pady=(10, 10))

entry_031F = Entry(frame_00F)
entry_031F.grid(row=1, column=3, pady=(10, 10), padx=(10, 10))
entry_031F.insert(0, 0)
label_04F = Label(frame_00F, text="Sheet 2:", bg="#004680",fg="WHITE",font=("Copperplate", 14))
label_04F.grid(row=1, column=4, padx=(10, 10), pady=(10, 10))

entry_04F = Entry(frame_00F)
entry_04F.grid(row=1, column=5, pady=(10, 10), padx=(10, 10))
entry_04F.insert(0, 0)
# left stationary

label_05F = Label(frame_00F, text="Files left:", bg="#004680",fg="WHITE",font=("Copperplate", 14))
label_05F.grid(row=2, column=0, padx=(10, 10), pady=(10, 10))

entry_05F = Entry(frame_00F)
entry_05F.grid(row=2, column=1, pady=(10, 10), padx=(10, 10))
entry_05F.insert(0, 100)
label_06F = Label(frame_00F, text="Sheet 1 Left:", bg="#004680",fg="WHITE",font=("Copperplate", 14) )
label_06F.grid(row=2, column=2, padx=(10, 10), pady=(10, 10))

entry_06F = Entry(frame_00F)
entry_06F.grid(row=2, column=3, pady=(10, 10), padx=(10, 10))
entry_06F.insert(0, 100)
label_07F = Label(frame_00F, text="Sheet 2 Left:", bg="#004680",fg="WHITE",font=("Copperplate", 14))
label_07F.grid(row=2, column=4, padx=(10, 10), pady=(10, 10))

entry_07F = Entry(frame_00F)
entry_07F.grid(row=2, column=5, pady=(10, 10), padx=(10, 10))
entry_07F.insert(0, 100)


button_05F = Button(frame_00F, text="PIE CHART",bg="#789FCF",fg="#000000",font=("Copperplate", 14))
button_05F.bind("<Button-1>", graph)
button_05F.grid(row=2, column=7, padx=(10, 10), pady=(10, 10))

root.mainloop()